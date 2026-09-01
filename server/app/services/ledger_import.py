"""台账导入（T5-4 / FR-D2）：Excel 解析 → 校验 → 落库。

与「在线编辑」不同，导入是**批量写库**，一旦写坏就是脏数据，
所以这里比导出保守得多：

    1. 表头必须是数据字典里登记过的中文列名，出现未知列直接整批拒绝
       （宁可让用户改表重传，也不猜列的含义）；
    2. 编码列接受「名称」或「编码」两种写法（人看的 Excel 通常写名称）；
    3. 类型逐格校验，遇到非法值报出**行号与列名**，而不是笼统的失败；
    4. 整批在一个事务里，任一行失败则全部回滚；
    5. 导入前后都落操作日志，便于事后追溯。

支持两种模式：append（追加，默认）与 replace（清空后导入）。
"""

from __future__ import annotations

import datetime
import io
from typing import Any

from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from ..core.exceptions import AppException, ErrorCode
from ..core.logging import get_logger
from . import ledger_service

logger = get_logger(__name__)

MAX_IMPORT_ROWS = 20000
MAX_ERRORS_REPORTED = 20


def parse_xlsx(content: bytes) -> tuple[list[str], list[list[Any]]]:
    """解析 Excel，返回 (表头, 数据行)。"""
    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise AppException(f"无法解析 Excel：{exc}", ErrorCode.BAD_REQUEST) from exc

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    rows = [r for r in rows if any(c is not None and str(c).strip() for c in r)]
    if not rows:
        raise AppException("Excel 内容为空", ErrorCode.BAD_REQUEST)

    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    return header, [list(r) for r in rows[1:]]


async def _name_to_code_map(
    ro: AsyncSession, col: dict
) -> dict[str, str]:
    """编码列的「名称 → 编码」映射，让用户可以在 Excel 里直接写名称。"""
    if not col.get("ref_table"):
        return {}
    rows = (
        await ro.execute(
            text(
                f"SELECT {col['ref_key']}, {col['ref_label']} FROM {col['ref_table']}"  # noqa: S608
            )
        )
    ).all()
    out: dict[str, str] = {}
    for code, name in rows:
        if name is None:
            continue
        out[str(name).strip()] = str(code)
        out[str(code)] = str(code)   # 也允许直接写编码
    return out


def _coerce(value: Any, col: dict, row_no: int, maps: dict[str, dict[str, str]]) -> Any:
    """把单元格值转成数据库可接受的类型；不合法则抛出带行号的错误。"""
    where = f"第 {row_no} 行「{col['cn_name']}」"
    name = col["column"]

    if value is None or (isinstance(value, str) and not value.strip()):
        return None

    if col["data_type"] == "number":
        if isinstance(value, bool):
            raise AppException(f"{where} 应是数值", ErrorCode.BAD_REQUEST)
        try:
            return float(value)
        except (TypeError, ValueError) as exc:
            raise AppException(f"{where} 不是合法数值：{value!r}", ErrorCode.BAD_REQUEST) from exc

    if col["data_type"] == "bool":
        if isinstance(value, bool):
            return value
        s = str(value).strip()
        if s in ("是", "true", "True", "1", "Y"):
            return True
        if s in ("否", "false", "False", "0", "N"):
            return False
        raise AppException(f"{where} 应是「是/否」：{value!r}", ErrorCode.BAD_REQUEST)

    if col["data_type"] == "date":
        if isinstance(value, datetime.datetime):
            return value.date()
        if isinstance(value, datetime.date):
            return value
        s = str(value).strip()
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.datetime.strptime(s, fmt).date()
            except ValueError:
                continue
        raise AppException(f"{where} 日期格式应为 YYYY-MM-DD：{s!r}", ErrorCode.BAD_REQUEST)

    # 文本 / 枚举：编码列先把名称翻成编码
    s = str(value).strip()
    mapping = maps.get(name)
    if mapping:
        if s not in mapping:
            raise AppException(
                f"{where} 找不到对应的编码：{s!r}（可填名称或编码）",
                ErrorCode.BAD_REQUEST,
            )
        return mapping[s]
    return s


async def import_rows(
    db: AsyncSession,
    ro: AsyncSession,
    key: str,
    content: bytes,
    *,
    mode: str = "append",
) -> dict:
    """导入并返回统计结果。任一行校验失败则整批回滚。"""
    header, raw = parse_xlsx(content)
    if not raw:
        raise AppException("Excel 没有数据行", ErrorCode.BAD_REQUEST)
    if len(raw) > MAX_IMPORT_ROWS:
        raise AppException(
            f"单次导入上限 {MAX_IMPORT_ROWS} 行，本次 {len(raw)} 行",
            ErrorCode.BAD_REQUEST,
        )

    meta = await ledger_service._meta(db, key)
    by_cn = {m["cn_name"]: m for m in meta}

    unknown = [h for h in header if h and h not in by_cn]
    if unknown:
        raise AppException(
            f"存在无法识别的列：{'、'.join(unknown[:5])}；"
            f"请改用「导出 CSV」得到的文件作为模板",
            ErrorCode.BAD_REQUEST,
            detail={"allowed": sorted(by_cn)},
        )

    cols = [by_cn[h] for h in header if h]
    if not cols:
        raise AppException("表头为空", ErrorCode.BAD_REQUEST)

    # 编码列的「名称 → 编码」映射
    maps: dict[str, dict[str, str]] = {}
    for m in cols:
        if m.get("ref_table"):
            maps[m["column"]] = await _name_to_code_map(ro, m)

    values: list[list[Any]] = []
    errors: list[str] = []
    for i, row in enumerate(raw, start=2):   # 第 1 行是表头
        try:
            values.append([
                _coerce(row[j] if j < len(row) else None, m, i, maps)
                for j, m in enumerate(cols)
            ])
        except AppException as exc:
            if len(errors) < MAX_ERRORS_REPORTED:
                errors.append(exc.message)

    if errors:
        raise AppException(
            f"校验未通过，共 {len(errors)} 处问题：{'；'.join(errors[:5])}",
            ErrorCode.BAD_REQUEST,
            detail={"errors": errors},
        )

    table = ledger_service._qualified(key)
    col_names = ", ".join(c["column"] for c in cols)
    placeholders = ", ".join(f":v{i}" for i in range(len(cols)))

    # AsyncSession 自带一个隐式事务，不能再 db.begin()（会报
    # "A transaction is already begun"）。这里直接执行后 commit，
    # 失败时 rollback——整批自然落在同一个事务里，不会写进一半。
    try:
        if mode == "replace":
            await db.execute(text(f"DELETE FROM {table}"))  # noqa: S608
        for vals in values:
            await db.execute(
                text(
                    f"INSERT INTO {table} ({col_names}) "  # noqa: S608
                    f"VALUES ({placeholders})"
                ),
                {f"v{i}": v for i, v in enumerate(vals)},
            )
        await db.commit()
    except Exception as exc:  # noqa: BLE001
        await db.rollback()
        logger.warning("台账导入失败，已回滚：%s", exc)
        raise AppException(
            f"写入数据库失败，已回滚：{exc}", ErrorCode.BAD_REQUEST
        ) from exc

    return {
        "imported": len(values),
        "columns": [c["column"] for c in cols],
        "mode": mode,
    }
